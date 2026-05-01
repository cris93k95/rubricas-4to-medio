"""
Rúbricas 4° Medio — Flask Backend
Evaluación Mock Job Interview (Video) + Curriculum Vitae
Persistencia en PostgreSQL · Autenticación Google OAuth
"""
from __future__ import annotations

import io
import json
import logging
import os
import threading
from copy import deepcopy
from functools import wraps
from pathlib import Path

from authlib.integrations.flask_client import OAuth
from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

BASE_DIR = Path(__file__).resolve().parent
DATABASE_URL = (os.getenv("DATABASE_URL") or "").strip()
SECRET_KEY = os.getenv("SECRET_KEY", "rubricas-dev-key-change-in-production")

GOOGLE_CLIENT_ID = (os.getenv("GOOGLE_CLIENT_ID") or "").strip()
GOOGLE_CLIENT_SECRET = (os.getenv("GOOGLE_CLIENT_SECRET") or "").strip()
GOOGLE_DISCOVERY_URL = "https://accounts.google.com/.well-known/openid-configuration"

ADMIN_EMAIL = (os.getenv("ADMIN_EMAIL") or "").strip().lower()
ADMIN_NAME = (os.getenv("ADMIN_NAME") or "Administrador").strip()

VALID_TOOLS = {"video", "cv", "shared"}

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
app.secret_key = SECRET_KEY
app.config['PREFERRED_URL_SCHEME'] = 'https'
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_HTTPONLY'] = True
lock = threading.Lock()
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

oauth = OAuth(app)
google_oauth_enabled = bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET)
if google_oauth_enabled:
    oauth.register(
        name="google",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        server_metadata_url=GOOGLE_DISCOVERY_URL,
        client_kwargs={"scope": "openid email profile"},
    )


# ─── Database ───────────────────────────────────────────────────────────────

def normalize_database_url(raw_url: str) -> str:
    if raw_url.startswith("postgres://"):
        return raw_url.replace("postgres://", "postgresql+psycopg://", 1)
    if raw_url.startswith("postgresql://") and "+" not in raw_url.split("://", 1)[0]:
        return raw_url.replace("postgresql://", "postgresql+psycopg://", 1)
    return raw_url


engine: Engine | None = None
if DATABASE_URL:
    engine = create_engine(normalize_database_url(DATABASE_URL), pool_pre_ping=True, future=True)


def default_state() -> dict:
    return {"courses": {}}


def init_database_if_needed() -> None:
    if not engine:
        return
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS app_users (
                id SERIAL PRIMARY KEY,
                email TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                role TEXT NOT NULL,
                google_sub TEXT,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                last_login TIMESTAMP
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS rubric_state (
                user_id INTEGER REFERENCES app_users(id) ON DELETE CASCADE,
                tool TEXT NOT NULL,
                data TEXT NOT NULL,
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (user_id, tool)
            )
        """))
        # Bootstrap admin user
        if ADMIN_EMAIL:
            logger.info(f"Bootstrapping admin user: email={ADMIN_EMAIL}, name={ADMIN_NAME}")
            existing = conn.execute(
                text("SELECT id FROM app_users WHERE email = :email"),
                {"email": ADMIN_EMAIL},
            ).first()
            if existing:
                conn.execute(
                    text("UPDATE app_users SET name = :name, role = 'admin' WHERE id = :id"),
                    {"id": existing[0], "name": ADMIN_NAME},
                )
                logger.info(f"Admin user updated (id={existing[0]})")
            else:
                conn.execute(
                    text("INSERT INTO app_users (email, name, role) VALUES (:email, :name, 'admin')"),
                    {"email": ADMIN_EMAIL, "name": ADMIN_NAME},
                )
                logger.info("Admin user CREATED")
        else:
            logger.warning("ADMIN_EMAIL not set! No admin user will be created.")


init_database_if_needed()


# ─── State Management ────────────────────────────────────────────────────────

def load_state(user_id: int, tool: str) -> dict:
    if engine:
        try:
            with engine.begin() as conn:
                row = conn.execute(
                    text("SELECT data FROM rubric_state WHERE user_id = :uid AND tool = :tool"),
                    {"uid": user_id, "tool": tool},
                ).first()
                if row and row[0]:
                    return json.loads(row[0])
        except Exception:
            pass
    return default_state()


def save_state(user_id: int, tool: str, state: dict) -> None:
    if engine:
        payload = json.dumps(state, ensure_ascii=False)
        with engine.begin() as conn:
            conn.execute(text("""
                INSERT INTO rubric_state (user_id, tool, data, updated_at)
                VALUES (:uid, :tool, :data, CURRENT_TIMESTAMP)
                ON CONFLICT (user_id, tool)
                DO UPDATE SET data = EXCLUDED.data, updated_at = CURRENT_TIMESTAMP
            """), {"uid": user_id, "tool": tool, "data": payload})


def pair_index(pair: dict | None) -> str | None:
    if not isinstance(pair, dict):
        return None
    pair_id = pair.get("id")
    if pair_id is None:
        return None
    return str(pair_id)


def merge_shared_pair(existing_pair: dict | None, incoming_pair: dict | None, source: str) -> dict:
    if not isinstance(existing_pair, dict):
        return deepcopy(incoming_pair) if isinstance(incoming_pair, dict) else {}
    if not isinstance(incoming_pair, dict):
        return deepcopy(existing_pair)

    if source == "video":
        merged = deepcopy(existing_pair)
        for key, value in incoming_pair.items():
            if key in {"cv_scores", "cv_feedback"}:
                continue
            merged[key] = deepcopy(value)
        return merged

    if source == "cv":
        merged = deepcopy(existing_pair)
        for key in ("cv_scores", "cv_feedback", "isOpen"):
            if key in incoming_pair:
                merged[key] = deepcopy(incoming_pair[key])
        return merged

    merged = deepcopy(existing_pair)
    for key, value in incoming_pair.items():
        merged[key] = deepcopy(value)
    return merged


def merge_shared_course(existing_course: dict | None, incoming_course: dict | None, source: str) -> dict:
    if not isinstance(existing_course, dict):
        return deepcopy(incoming_course) if isinstance(incoming_course, dict) else {}
    if not isinstance(incoming_course, dict):
        return deepcopy(existing_course)

    if source == "video":
        merged = deepcopy(existing_course)
        for key, value in incoming_course.items():
            if key == "pairs":
                continue
            merged[key] = deepcopy(value)

        existing_pairs = {
            pair_index(pair): pair
            for pair in existing_course.get("pairs", [])
            if pair_index(pair) is not None
        }
        merged["pairs"] = []
        for incoming_pair in incoming_course.get("pairs", []):
            incoming_key = pair_index(incoming_pair)
            if incoming_key in existing_pairs:
                merged["pairs"].append(merge_shared_pair(existing_pairs[incoming_key], incoming_pair, source))
            else:
                merged["pairs"].append(deepcopy(incoming_pair))
        return merged

    if source == "cv":
        merged = deepcopy(existing_course)
        if "isOpen" in incoming_course:
            merged["isOpen"] = deepcopy(incoming_course["isOpen"])

        existing_pairs = existing_course.get("pairs", [])
        incoming_pairs = {
            pair_index(pair): pair
            for pair in incoming_course.get("pairs", [])
            if pair_index(pair) is not None
        }

        if not existing_pairs:
            merged["pairs"] = [deepcopy(pair) for pair in incoming_course.get("pairs", [])]
            return merged

        merged_pairs = []
        for existing_pair in existing_pairs:
            existing_key = pair_index(existing_pair)
            if existing_key is not None and existing_key in incoming_pairs:
                merged_pairs.append(merge_shared_pair(existing_pair, incoming_pairs[existing_key], source))
            else:
                merged_pairs.append(deepcopy(existing_pair))
        merged["pairs"] = merged_pairs
        return merged

    merged = deepcopy(existing_course)
    for key, value in incoming_course.items():
        merged[key] = deepcopy(value)
    return merged


def merge_shared_state(current_state: dict | None, incoming_state: dict | None, source: str) -> dict:
    current_courses = current_state.get("courses", {}) if isinstance(current_state, dict) else {}
    incoming_courses = incoming_state.get("courses", {}) if isinstance(incoming_state, dict) else {}

    if source == "video":
        merged_courses = {}
        for course_name, incoming_course in incoming_courses.items():
            merged_courses[course_name] = merge_shared_course(current_courses.get(course_name), incoming_course, source)
        return {"courses": merged_courses}

    if source == "cv":
        merged_courses = {course_name: deepcopy(course_state) for course_name, course_state in current_courses.items()}
        for course_name, incoming_course in incoming_courses.items():
            if course_name in current_courses:
                merged_courses[course_name] = merge_shared_course(current_courses[course_name], incoming_course, source)
            else:
                merged_courses[course_name] = deepcopy(incoming_course)
        return {"courses": merged_courses}

    return deepcopy(incoming_state) if isinstance(incoming_state, dict) else default_state()


# ─── Auth Helpers ─────────────────────────────────────────────────────────────

def normalize_email(email: str) -> str:
    return (email or "").strip().lower()


def get_user_by_id(user_id: int) -> dict | None:
    if not engine:
        return None
    with engine.begin() as conn:
        row = conn.execute(
            text("SELECT id, email, name, role FROM app_users WHERE id = :id"),
            {"id": user_id},
        ).first()
    if not row:
        return None
    return {"id": int(row[0]), "email": row[1], "name": row[2], "role": row[3]}


def get_user_by_email(email: str) -> dict | None:
    if not engine:
        return None
    with engine.begin() as conn:
        row = conn.execute(
            text("SELECT id, email, name, role FROM app_users WHERE email = :email"),
            {"email": normalize_email(email)},
        ).first()
    if not row:
        return None
    return {"id": int(row[0]), "email": row[1], "name": row[2], "role": row[3]}


def get_current_user() -> dict | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    try:
        return get_user_by_id(int(user_id))
    except Exception:
        return None


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not get_current_user():
            if request.path.startswith("/api"):
                return jsonify({"error": "No autenticado"}), 401
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


# ─── Auth Routes ──────────────────────────────────────────────────────────────

@app.route("/login")
def login():
    user = get_current_user()
    if user:
        return redirect(url_for("landing"))
    return render_template("login.html", google_enabled=google_oauth_enabled)


@app.route("/auth/google")
def auth_google():
    if not google_oauth_enabled:
        return "Google OAuth no configurado", 503
    redirect_uri = url_for("auth_google_callback", _external=True, _scheme='https')
    logger.info(f"OAuth redirect URI: {redirect_uri}")
    return oauth.google.authorize_redirect(redirect_uri)


@app.route("/auth/google/callback")
def auth_google_callback():
    if not google_oauth_enabled:
        return "Google OAuth no configurado", 503
    try:
        token = oauth.google.authorize_access_token()
        user_info = token.get("userinfo")
        if not user_info:
            resp = oauth.google.get("userinfo")
            user_info = resp.json()
    except Exception as e:
        logger.error(f"OAuth error: {e}", exc_info=True)
        return f"Error al autenticar con Google: {e}", 401

    email = normalize_email(user_info.get("email", ""))
    logger.info(f"Google login attempt: email={email}")
    if not email:
        return "No se pudo obtener email de Google", 401

    user = get_user_by_email(email)
    if not user:
        # Auto-create admin if email matches
        if email == ADMIN_EMAIL and engine:
            try:
                logger.info(f"Auto-creating admin user: {email}")
                with engine.begin() as conn:
                    # Ensure tables exist
                    conn.execute(text("""
                        CREATE TABLE IF NOT EXISTS app_users (
                            id SERIAL PRIMARY KEY,
                            email TEXT NOT NULL UNIQUE,
                            name TEXT NOT NULL,
                            role TEXT NOT NULL,
                            google_sub TEXT,
                            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                            last_login TIMESTAMP
                        )
                    """))
                    conn.execute(text("""
                        CREATE TABLE IF NOT EXISTS rubric_state (
                            user_id INTEGER REFERENCES app_users(id) ON DELETE CASCADE,
                            tool TEXT NOT NULL,
                            data TEXT NOT NULL,
                            updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                            PRIMARY KEY (user_id, tool)
                        )
                    """))
                    conn.execute(text("""
                        INSERT INTO app_users (email, name, role)
                        VALUES (:email, :name, 'admin')
                        ON CONFLICT (email) DO NOTHING
                    """), {"email": email, "name": ADMIN_NAME})
                user = get_user_by_email(email)
                logger.info(f"Admin user created successfully: {user}")
            except Exception as e:
                logger.error(f"Failed to auto-create admin: {e}", exc_info=True)
                return f"Error creando usuario admin: {e}", 500
        if not user:
            logger.warning(f"User NOT FOUND in DB: {email}. ADMIN_EMAIL={ADMIN_EMAIL}")
            return f"Usuario no autorizado ({email}).", 403

    if engine:
        with engine.begin() as conn:
            conn.execute(
                text("UPDATE app_users SET google_sub = :sub, last_login = CURRENT_TIMESTAMP WHERE id = :id"),
                {"id": user["id"], "sub": user_info.get("sub")},
            )

    session["user_id"] = int(user["id"])
    return redirect(url_for("landing"))


@app.route("/logout")
def logout():
    session.pop("user_id", None)
    return redirect(url_for("login"))


# ─── Page Routes ──────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def landing():
    return render_template("landing.html", current_user=get_current_user())


@app.route("/video")
@login_required
def video_page():
    return render_template("video.html", current_user=get_current_user())


@app.route("/cv")
@login_required
def cv_page():
    return render_template("cv.html", current_user=get_current_user())


# ─── API: State ───────────────────────────────────────────────────────────────

@app.route("/api/state/<tool>", methods=["GET"])
@login_required
def api_get_state(tool: str):
    if tool not in VALID_TOOLS:
        return jsonify({"error": "Herramienta no válida"}), 400
    user = get_current_user()
    with lock:
        state = load_state(int(user["id"]), tool)
    return jsonify(state)


@app.route("/api/state/<tool>", methods=["PUT"])
@login_required
def api_save_state(tool: str):
    if tool not in VALID_TOOLS:
        return jsonify({"error": "Herramienta no válida"}), 400
    user = get_current_user()
    payload = request.get_json(force=True)
    source = (request.headers.get("X-Rubrica-Source") or "").strip().lower()
    with lock:
        if tool == "shared" and source in {"video", "cv"}:
            current_state = load_state(int(user["id"]), tool)
            payload = merge_shared_state(current_state, payload, source)
        save_state(int(user["id"]), tool, payload)
    return jsonify({"ok": True})


# ─── API: Excel Upload ───────────────────────────────────────────────────────

@app.route("/api/<tool>/upload-excel/<course_name>", methods=["POST"])
@login_required
def api_upload_excel(tool: str, course_name: str):
    if tool not in VALID_TOOLS:
        return jsonify({"error": "Herramienta no válida"}), 400
    
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Archivo requerido"}), 400

    try:
        wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
        ws = wb[wb.sheetnames[0]]
        names = []
        for row in ws.iter_rows(min_row=1, max_col=3):
            values = [str(cell.value).strip() for cell in row if cell.value is not None]
            # Ignore headers and numbers
            name_parts = [v for v in values if v and not v.isdigit() and len(v) > 1 and v.lower() not in ('nombre', 'apellido', 'nombres', 'apellidos', 'alumno', 'alumnos', 'estudiante', 'estudiantes')]
            if name_parts:
                names.append(" ".join(name_parts))
    except Exception as e:
        return jsonify({"error": f"Error al leer archivo: {e}"}), 400

    user = get_current_user()
    with lock:
        state = load_state(int(user["id"]), tool)
        courses = state.get("courses", {})
        
        if course_name not in courses:
            return jsonify({"error": "Curso no encontrado"}), 404

        added = 0
        if tool == "shared" or tool == "video":
            # Add to roster
            roster = courses[course_name].get("roster", [])
            for n in names:
                if n not in roster:
                    roster.append(n)
                    added += 1
            courses[course_name]["roster"] = roster

        state["courses"] = courses
        save_state(int(user["id"]), tool, state)

    return jsonify({"ok": True, "added": added})


# ─── API: Export / Import ─────────────────────────────────────────────────────

@app.route("/api/export", methods=["GET"])
@login_required
def api_export():
    user = get_current_user()
    with lock:
        video_state = load_state(int(user["id"]), "video")
        cv_state = load_state(int(user["id"]), "cv")
    
    export_data = {"video": video_state, "cv": cv_state}
    payload = json.dumps(export_data, ensure_ascii=False, indent=2).encode("utf-8")
    
    from datetime import datetime
    return send_file(
        io.BytesIO(payload),
        as_attachment=True,
        download_name=f"respaldo_rubricas_4to_{datetime.now().date()}.json",
        mimetype="application/json",
    )


@app.route("/api/import", methods=["POST"])
@login_required
def api_import():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Archivo requerido"}), 400
    try:
        imported = json.loads(file.read().decode("utf-8"))
    except Exception:
        return jsonify({"error": "JSON inválido"}), 400

    user = get_current_user()
    with lock:
        if "video" in imported:
            save_state(int(user["id"]), "video", imported["video"])
        if "cv" in imported:
            save_state(int(user["id"]), "cv", imported["cv"])
    return jsonify({"ok": True})


# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
